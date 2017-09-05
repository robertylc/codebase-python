#!/usr/local/bin/python3
# coding=utf-8
from openpyxl import load_workbook
from openpyxl import Workbook

def main():
    ## 创建excel文件、sheet
    wb = Workbook()
    ws = wb.active()      #默认创建的sheet
    ws1 = wb.create_sheet(0)
    ws = wb['worksheet1']  # 选中工作表

    ## sheet属性操作
    ws.title = "name";

    ## 单元操作
    c = ws['A4']
    ws['A4'] = 4
    d = ws.cell(row=4, column=2)
    cell_range = ws['A1':'C2']

    ## excel打开、保存
    wb = load_workbook(filename="sample.xlsx")
    wb.save("output_sample.xlsx")

if __name__ == '__main__':
    #main()
    wb = Workbook()
    ws = wb.create_sheet(title="test")
    ws['A4'] = "test"
    d = ws.cell(row=4, column=2)
    d.value = "test2"
    wb.save(r'''E:\PycharmProjects\codebase\result.xlsx''')

